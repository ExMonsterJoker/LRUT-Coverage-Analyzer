import os
import glob
import pandas as pd
import openpyxl
import re

#test

INPUT_DIR = 'Input'
OUTPUT_SUFFIX = '_coverage.xlsx'

TARGET_HEADERS = [
    'No', 'PKM', 'Test Point', 'Number Of Anomaly', 
    'Length of Test (M)', 'North', 'East', 'From Datum (M)', 
    't reading of Thinning / Anomaly, mm', 'Remark'
]

def convert_dms_to_dd(dms_value):
    """
    Converts DMS string (e.g. 12°34'56"N) to Decimal Degrees.
    Handles various separators by treating non-numeric characters as delimiters.
    """
    if dms_value is None:
        return None
    
    # If it's already a number, return it (assuming it's already DD)
    if isinstance(dms_value, (int, float)):
        return float(dms_value)
        
    dms_str = str(dms_value).strip().upper()
    if not dms_str:
        return None
        
    # Check direction for sign
    is_negative = False
    if 'S' in dms_str or 'W' in dms_str:
        is_negative = True
        
    # Replace non-digit/non-decimal characters with space
    # This handles degree symbols, minutes, seconds symbols, etc.
    cleaned = re.sub(r'[^0-9.]', ' ', dms_str)
    parts = [float(x) for x in cleaned.split() if x]
    
    dd = 0.0
    if len(parts) >= 3:
        # Degrees, Minutes, Seconds
        dd = parts[0] + parts[1]/60 + parts[2]/3600
    elif len(parts) == 2:
        # Degrees, Decimal Minutes
        dd = parts[0] + parts[1]/60
    elif len(parts) == 1:
        # Decimal Degrees
        dd = parts[0]
        
    if is_negative:
        dd = -dd
        
    return dd

def find_headers(sheet):
    """
    Scans columns A-AZ (1-52) and Rows 1-10 to find headers.
    Returns a dict {header_name: column_index (0-based)}.
    Also returns the row_index where headers were found (assuming all on same row or we take the max?)
    The prompt implies they might be scattered, but usually headers are on one row. 
    However, "Start scanning... Find all header" suggests searching. 
    We will store (row_idx, col_idx) for each header.
    """
    found_headers = {}
    
    # Scan Row 1 to 10 (0 to 9 in 0-based index)
    # Scan Column A to AZ (1 to 52)
    
    for r in range(0, 10):
        for c in range(1, 53): # A is 1, AZ is 52
            cell_val = sheet.cell(row=r+1, column=c).value
            if cell_val:
                s_val = str(cell_val).strip()
                s_val_lower = s_val.lower()
                for target in TARGET_HEADERS:
                    if target == 't reading of Thinning / Anomaly, mm':
                        if 't reading of thinning' in s_val_lower or 't reading o thinning' in s_val_lower or 't actual anomaly, mm' in s_val_lower or 't reading of thinning / anomaly, mm' in s_val_lower:
                            found_headers[target] = {'row': r, 'col': c}
                            break
                    else:
                        if s_val_lower == target.lower():
                            found_headers[target] = {'row': r, 'col': c}
                            break
                    
    return found_headers

def process_file(filepath):
    print(f"Processing {filepath}...")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Error opening {filepath}: {e}")
        return

    # Case insensitive sheet search
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == 'recommendation':
            target_sheet = wb[sheet_name]
            break
            
    if not target_sheet:
        print(f"Sheet 'Recommendation' not found in {filepath}")
        return

    # 1. Find Headers
    headers_map = find_headers(target_sheet)
    
    missing = [h for h in TARGET_HEADERS if h not in headers_map]
    if missing:
        print(f"Warning: Missing headers in {filepath}: {missing}")
        # Proceeding might be dangerous if critical headers are missing, but let's try.
        if 'No' not in headers_map:
            print("Critical header 'No' missing. Skipping.")
            return

    # 2. Find Start and End Rows based on 'No' column
    no_col_idx = headers_map['No']['col'] # 1-based index from openpyxl
    no_header_row = headers_map['No']['row'] + 1 # 1-based row
    
    start_row = -1
    
    # Scan for '1' in 'No' column
    # We start searching below the header
    current_row = no_header_row + 1
    max_scan_row = 5000 # Safety breakdown
    
    while current_row < max_scan_row:
        val = target_sheet.cell(row=current_row, column=no_col_idx).value
        # Check if value is 1 (integer or string "1")
        if val == 1 or str(val).strip() == '1':
            start_row = current_row
            break
        current_row += 1
        
    if start_row == -1:
        print("Could not find start row (value '1' in 'No' column).")
        return

    # Scan for End Row
    # "Keep scanning until finding a cell with value of non-number 10 times. Keep last cell that has number value"
    end_row = start_row
    non_number_buffer_count = 0
    scan_ptr = start_row
    
    while non_number_buffer_count < 10 and scan_ptr < max_scan_row:
        val = target_sheet.cell(row=scan_ptr, column=no_col_idx).value
        is_number = False
        if val is not None:
            try:
                float(val)
                is_number = True
            except ValueError:
                pass
        
        if is_number:
            end_row = scan_ptr
            non_number_buffer_count = 0
        else:
            non_number_buffer_count += 1
            
        scan_ptr += 1

    print(f"Data range: Rows {start_row} to {end_row}")

    # 3. Extract Data
    data_rows = []
    
    for r in range(start_row, end_row + 1):
        row_data = {'_row_idx': r} # Keep track of original row index
        for header in TARGET_HEADERS:
            if header in headers_map:
                col = headers_map[header]['col']
                val = target_sheet.cell(row=r, column=col).value
                
                # Convert DMS to DD for North and East columns
                if header in ['North', 'East']:
                    val = convert_dms_to_dd(val)
                    
                row_data[header] = val
            else:
                row_data[header] = None
        
        # Special Extraction for TP row columns (BWD, FWD)
        # Even if not a TP row, we don't know yet, so let's defer reading BWD/FWD until we identify TP rows
        # But we need to read them relative to extraction time or keep the sheet open.
        # Let's read the potential BWD/FWD values just in case?
        # Or better: Identify TP rows during this pass or iterate logic after.
        # The prompt says: "When extracting data from column 'Number Of Anomaly', you will find row where has the value of 'TP'"
        
        # Check for TP here to grab extra columns
        num_anomaly_val = row_data.get('Number Of Anomaly')
        if num_anomaly_val == 'TP':
            # Extract next column and next next column relative to 'Number Of Anomaly' header col
            if 'Number Of Anomaly' in headers_map:
                base_col = headers_map['Number Of Anomaly']['col']
                bwd_val = target_sheet.cell(row=r, column=base_col + 1).value
                fwd_val = target_sheet.cell(row=r, column=base_col + 2).value
                row_data['_BWD_RAW'] = bwd_val
                row_data['_FWD_RAW'] = fwd_val
        
        data_rows.append(row_data)

    df = pd.DataFrame(data_rows)
    
    # 4. Transformations
    
    # Grouping logic
    # First, forward fill 'Test Point' if it's sparse? 
    # "You will find group row where it has same 'Test Point' value. Group them."
    # If the file has merged cells or empty cells for Test Point, standard pandas ffill is usually required.
    # Looking at the requirement: "found TP at row 12, and row 12 is inside group 2. then it is test point 2."
    # This implies we rely on row proximity (grouping).
    
    if 'Test Point' in df.columns:
        df['Test Point'] = df['Test Point'].ffill()
    
    if df.empty:
        print("No data extracted.")
        return

    # Identify groups
    # We accept that the data is sorted by Test Point roughly or grouped physically.
    groups = df.groupby('Test Point')
    
    final_rows = []
    
    # Store processed groups to do gap analysis later
    processed_groups = []

    for tp_name, group in groups:
        # Convert group to records
        group_records = group.to_dict('records')
        
        tp_row = None
        min_thickness = float('inf')
        thickness_vals = []
        valid_group_rows = []

        for row in group_records:
            if row.get('Number Of Anomaly') == 'TP':
                tp_row = row
            else:
                valid_group_rows.append(row)
                t_val = row.get('t reading of Thinning / Anomaly, mm')
                if t_val is not None:
                     try:
                         f_val = float(t_val)
                         thickness_vals.append(f_val)
                     except:
                         pass

        if thickness_vals:
            min_t = min(thickness_vals)
        else:
            min_t = None

        final_group_rows = []
        final_group_rows.extend(valid_group_rows)
        
        if tp_row:
            try:
                pkm_base = tp_row.get('PKM')
                bwd_raw = tp_row.get('_BWD_RAW')
                fwd_raw = tp_row.get('_FWD_RAW')
                
                pkm_base = float(pkm_base) if pkm_base is not None else 0.0
                bwd_val = float(bwd_raw) if bwd_raw is not None else 0.0
                fwd_val = float(fwd_raw) if fwd_raw is not None else 0.0
                
                # Backward Row
                # 'Number Of Anomaly' = "BWD"
                bwd_row = tp_row.copy()
                bwd_row['PKM'] = pkm_base - bwd_val
                bwd_row['Test Point'] = tp_name
                bwd_row['t reading of Thinning / Anomaly, mm'] = min_t
                bwd_row['Number Of Anomaly'] = "BWD"
                
                # Forward Row
                # 'Number Of Anomaly' = "FWD"
                fwd_row = tp_row.copy()
                fwd_row['PKM'] = pkm_base + fwd_val
                fwd_row['Test Point'] = tp_name
                fwd_row['t reading of Thinning / Anomaly, mm'] = min_t
                fwd_row['Number Of Anomaly'] = "FWD" 
                
                # Cleanup internal keys
                for r in [bwd_row, fwd_row]:
                    if '_BWD_RAW' in r: del r['_BWD_RAW']
                    if '_FWD_RAW' in r: del r['_FWD_RAW']
                    if '_row_idx' in r: del r['_row_idx']

                final_group_rows.append(tp_row)
                final_group_rows.append(bwd_row)
                final_group_rows.append(fwd_row)
                
            except Exception as e:
                print(f"Error creating coverage rows for Group {tp_name}: {e}")
        
        # Calculate Start PKM Group and End PKM Group
        # "Start PKM Group = minimum PKM in the group"
        # "END PKM Group = Maximum PKM in the group"
        group_pkms = []
        for r in final_group_rows:
            p = r.get('PKM')
            if p is not None:
                try:
                    group_pkms.append(float(p))
                except:
                    pass
        
        start_pkm_group = min(group_pkms) if group_pkms else None
        end_pkm_group = max(group_pkms) if group_pkms else None
        
        # Update rows with new columns
        for r in final_group_rows:
            r['Start PKM Group'] = start_pkm_group
            r['End PKM Group'] = end_pkm_group
            
            # Clean up raw keys from valid_group_rows if they exist
            if '_BWD_RAW' in r: del r['_BWD_RAW']
            if '_FWD_RAW' in r: del r['_FWD_RAW']
            if '_row_idx' in r: del r['_row_idx']

        processed_groups.append({
            'start_pkm': start_pkm_group if start_pkm_group is not None else -1.0,
            'end_pkm': end_pkm_group if end_pkm_group is not None else -1.0,
            'rows': final_group_rows
        })

    # Sort groups based on Start PKM Group
    processed_groups.sort(key=lambda x: x['start_pkm'])
    
    # Gap Analysis
    # "If there is gap PKM between group, create a new row... Start PKM Group = Maximum PKM in previous group..."
    final_output_rows = []
    
    for i in range(len(processed_groups)):
        current_group = processed_groups[i]
        final_output_rows.extend(current_group['rows'])
        
        if i < len(processed_groups) - 1:
            next_group = processed_groups[i+1]
            
            gap_start = current_group['end_pkm']
            gap_end = next_group['start_pkm']
            
            if gap_end > gap_start:
                # Create Gap Row
                gap_row = {}
                # Initialize with None for all headers to be safe
                for h in TARGET_HEADERS:
                    gap_row[h] = None
                    
                gap_row['PKM'] = gap_start # Position it at start of gap
                gap_row['Start PKM Group'] = gap_start
                gap_row['End PKM Group'] = gap_end
                gap_row['Test Point'] = "GAP" # Marker
                gap_row['Number Of Anomaly'] = "GAP" # Optional clarity
                
                final_output_rows.append(gap_row)
    
    result_df = pd.DataFrame(final_output_rows)
    
    # 5. Sort
    # "dump column dataframe with two sorts steps... Sort by start PKM group... sort by PKM group"
    # Actually 'PKM Group' probably refers to 'PKM' value? Or 'Start PKM Group'?
    # "Sort by start PKM group... sort by PKM group" -> Sort by Start PKM Group, then by PKM.
    
    if result_df.empty:
        print("No data to save.")
        return

    # Ensure columns exist even if empty (for gap rows)
    if 'Start PKM Group' not in result_df.columns: result_df['Start PKM Group'] = None
    if 'PKM' not in result_df.columns: result_df['PKM'] = None
    
    result_df['Start PKM Group'] = pd.to_numeric(result_df['Start PKM Group'], errors='coerce')
    result_df['PKM'] = pd.to_numeric(result_df['PKM'], errors='coerce')
    
    result_df = result_df.sort_values(by=['Start PKM Group', 'PKM'])
        
    # Export
    base, ext = os.path.splitext(filepath)
    new_filename = f"{base}{OUTPUT_SUFFIX}"
    
    print(f"Saving to {new_filename}")
    result_df.to_excel(new_filename, index=False)


def main():
    files = glob.glob(os.path.join(INPUT_DIR, '*.xlsx'))
    files.extend(glob.glob(os.path.join(INPUT_DIR, '*.xls')))
    
    print(f"Found {len(files)} files in {INPUT_DIR}")
    
    for f in files:
        if not f.endswith(OUTPUT_SUFFIX): # Avoid processing output files if they exist
            process_file(f)

if __name__ == "__main__":
    main()
